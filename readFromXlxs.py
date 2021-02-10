import time

from selenium import webdriver
from openpyxl import load_workbook

count = 1
file_name = '100000.xlsx'
path = r'C:\Users\Laptop\Desktop\\' + file_name

if __name__ == '__main__':
    excel_document = load_workbook(path)
    web = webdriver.Firefox()
    web.get("https://www.elections.ps/tabid/596/language/ar-PS/Default.aspx")
    sheet = excel_document['ورقة1']
    sleepAfter = 1
    for i in range(1, sheet.max_row + 1):
        if sleepAfter > 20:
            time.sleep(30)
            sleepAfter = 1
        sleepAfter += 1
        a = 'A' + str(i)
        b = 'B' + str(i)
        c = 'C' + str(i)
        id = sheet[a].value
        yob = sheet[b].value

        web.find_element_by_id("dnn_ctr4525_View_PalID").send_keys(id)
        web.find_element_by_id("dnn_ctr4525_View_YearOfBirth").send_keys(yob)
        web.find_element_by_id('dnn_ctr4525_View_recaptcha_demo_submit').click()
        while True:
            try:
                print(str(count) + '-' + str(id) + ' : ' + web.find_element_by_id('dnn_ctr4525_ResultView_lblCenter').text)
                sheet[c].value = web.find_element_by_id('dnn_ctr4525_ResultView_lblCenter').text
                count += 1
                break
            except:
                True
            try:
                web.find_element_by_id('dnn_ctr1204_VoterSearchAR_DetailsView1_Empt')
                print(str(count) + '-' + str(id) + ' : ' + 'غير مسجل')
                sheet[c].value = 'غير مسجل'
                count += 1
                break
            except:
                True

        excel_document.save(file_name)
        web.back();

    web.close()
