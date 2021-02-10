from openpyxl import load_workbook

count = 1
file_name ='60000_1.xlsx'
path = r'C:\Users\Laptop\Desktop\\' + file_name
if __name__ == '__main__':
    excel_document = load_workbook(path)
    sheet = excel_document['ورقة1']
    for i in range(1, sheet.max_row + 1):
        B = 'B' + str(i)
        C = 'C' + str(i)
        try:
            yob = int(str(sheet[B].value).split('/')[2])
        except:
            try:
                yob = int(str(sheet[B].value).split()[0].split('-')[0])
            except:
                continue
        sheet[C].value = yob
    excel_document.save(file_name)



